from __future__ import annotations

"""
엔트로피 기반 유속분포법 - 외부 Excel 입력 버전

입력 파일
1) 횡단면 Excel
   - 기본 파일명: XS_sample.xlsx
   - 기본 열: x(m), y(EL.m)

2) 표면유속 Excel
   - 기본 파일명: Vel_sample.xlsx
   - 기본 열: 측선번호, 측선거리(m), 유속(m/s)
   - 행의 수가 곧 측선 수이므로 측선 증감에 자동 대응

주요 출력
- vpm_output/isovelocity_contour.png
- vpm_output/water_level_area_table.xlsx
- vpm_output/velocity_grid.csv
- vpm_output/measurement_fit.csv
- vpm_output/selected_channel_cross_section.csv
- vpm_output/calculation_summary.json

Umax 처리
- 관측 최대표면유속과 비제약 적합 속도배율을 분리
- 기본적으로 Model Umax >= 관측 최대표면유속 제약 적용
- max_velocity_depth_ratio로 최대유속의 수면 아래 위치 설정
  · 0.00: 수면
  · 0.05: 국부수심의 5% 아래

저수위 측선 처리
- 현재 수위에서 선택된 주수로 밖의 측선은 기본적으로 자동 제외
- 제외 측선은 실행창 경고, measurement_fit.csv 제외사유,
  calculation_summary.json에 기록
- 선택 주수로 안의 양의 유속 측선이 최소 개수보다 적으면 계산 중단

실행 예
    python velocity_distribution_excel_input.py

별도 경로나 수위를 지정할 때
    python velocity_distribution_excel_input.py ^
        --xs XS_sample.xlsx ^
        --vel Vel_sample.xlsx ^
        --water-level 4.50

그래프 창을 띄우지 않을 때
    python velocity_distribution_excel_input.py --no-show

필요 패키지
    pip install numpy pandas matplotlib openpyxl
"""

from dataclasses import dataclass, replace
from pathlib import Path
from typing import Iterable
import argparse
import json
import math
import re
import sys

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# 1. 설정
# =============================================================================

BASE_DIR = Path(__file__).resolve().parent


@dataclass
class Measurement:
    x: float
    surface_velocity: float
    name: str = ""
    quality_weight: float = 1.0
    source_row: int | None = None


@dataclass
class ModelConfig:
    # 입력 파일
    cross_section_file: Path = BASE_DIR / "평창교_xs.xlsx"
    velocity_file: Path = BASE_DIR / "평창교_vel.xlsx"
    cross_section_sheet: str | int = 0
    velocity_sheet: str | int = 0

    # 실시간 수위 El.m
    water_level: float = 287.412 + 1.63

    # 격자 계산조건
    horizontal_grid_m: float = 1.00
    vertical_grid_m: float = 0.25       #default : 0.25

    # 엔트로피 매개변수
    # entropy_M=None이면 phi로부터 M을 자동 역산
    phi: float = 0.667      #default : 0.667
    entropy_M: float | None = None

    # 등유속선 형상 매개변수
    beta_left: float = 1.0
    beta_right: float = 1.0
    # 최대유속 발생 위치를 "수면 아래 깊이 / 국부수심"의 비율로 설정
    # 0.00: 최대유속이 수면에서 발생
    # 0.05: 각 위치에서 수면 아래 국부수심의 5% 지점에서 발생
    # 허용범위: 0.0 이상 1.0 미만
    max_velocity_depth_ratio: float = 0.00
    delta_y: float = 0.0
    delta_z_left: float = 0.0
    delta_z_right: float = 0.0

    # 여러 측선 자료로 공통 beta 자동 적합
    auto_fit_common_beta: bool = True
    beta_search_min: float = 0.10
    beta_search_max: float = 80.0
    beta_search_steps: int = 1200

    # flow_center_x=None일 때 측정 표면유속 전체를 이용하여
    # 흐름 중심 x, beta_left, beta_right를 함께 자동 적합한다.
    # 측선이 부족하면 기존 공통 beta 적합으로 자동 전환한다.
    auto_fit_flow_center_asymmetric_beta: bool = True
    joint_fit_center_search_steps: int = 41
    joint_fit_beta_search_steps: int = 31
    joint_fit_refinement_rounds: int = 2

    # None이면 최심점 x를 초기값으로 두고, 위 설정에 따라 자동 공동 적합
    flow_center_x: float | None = None

    # least_squares / median / mean
    max_velocity_method: str = "least_squares"

    # 물리 제약: 단면 최대유속은 관측 최대표면유속보다 작을 수 없음
    # True이면 Umax >= max(측정 표면유속)을 강제한다.
    enforce_umax_ge_observed_surface: bool = True

    # 유속 0 자료도 적합에 포함할지 여부
    include_zero_velocity_in_fit: bool = True

    # 현재 수위에서 선택 주수로 밖에 있는 측선 처리
    # "exclude": 경고 및 제외사유를 남기고 계산 계속
    # "error": 기존처럼 오류 발생
    outside_wetted_measurement_policy: str = "exclude"

    # 계산을 계속하기 위해 필요한 선택 주수로 내 양의 유속 측선 최소 개수
    minimum_positive_measurements: int = 2

    # 수위-단면적표 조건
    # None이면 최저 하상표고에서 현재 수위까지 생성
    area_table_min_level: float | None = None
    area_table_max_level: float | None = None
    area_table_interval_m: float = 0.10

    # 주수로 선택 기준 x
    # None이면 횡단면 전체의 최심점 x를 주수로 기준으로 사용
    area_reference_x: float | None = None

    # 출력
    contour_level_count: int = 16
    output_directory: str = "vpm_output_excel_v3"
    show_plot: bool = True


CONFIG = ModelConfig()


# =============================================================================
# 2. Excel 입력
# =============================================================================

def normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.strip().lower()
    return re.sub(r"[\s_\-./()\[\]{}]", "", text)


def find_column(
    frame: pd.DataFrame,
    aliases: set[str],
    *,
    fallback_index: int | None = None,
) -> object:
    normalized = {normalize_header(column): column for column in frame.columns}
    normalized_aliases = {normalize_header(alias) for alias in aliases}

    for key, original in normalized.items():
        if key in normalized_aliases:
            return original

    if fallback_index is not None and len(frame.columns) > fallback_index:
        return frame.columns[fallback_index]

    raise ValueError(
        "필요한 열을 찾을 수 없습니다. "
        f"현재 열={list(frame.columns)}, 허용 이름={sorted(aliases)}"
    )


def load_cross_section_excel(
    path: Path,
    sheet_name: str | int = 0,
) -> np.ndarray:
    if not path.exists():
        raise FileNotFoundError(f"횡단면 파일을 찾을 수 없습니다: {path}")

    frame = pd.read_excel(path, sheet_name=sheet_name)

    x_column = find_column(
        frame,
        {
            "x",
            "x(m)",
            "횡단거리",
            "횡단거리(m)",
            "거리",
            "station",
            "chainage",
        },
        fallback_index=0,
    )
    y_column = find_column(
        frame,
        {
            "y",
            "y(el.m)",
            "y(elm)",
            "하상표고",
            "하상표고(m)",
            "표고",
            "elevation",
            "el",
        },
        fallback_index=1,
    )

    data = frame[[x_column, y_column]].copy()
    data.columns = ["x", "bed_elevation"]
    data["x"] = pd.to_numeric(data["x"], errors="coerce")
    data["bed_elevation"] = pd.to_numeric(
        data["bed_elevation"], errors="coerce"
    )
    data = data.dropna()

    section = data.to_numpy(dtype=float)

    if len(section) < 2:
        raise ValueError("유효한 횡단면 좌표가 2개 미만입니다.")

    # 중복 x는 수직벽 표현일 수 있으므로 삭제하지 않는다.
    if np.any(np.diff(section[:, 0]) < 0):
        raise ValueError(
            "횡단거리 x가 감소하는 구간이 있습니다. "
            "Excel의 행 순서를 좌안→우안 방향으로 정렬해 주세요."
        )

    return section


def load_velocity_excel(
    path: Path,
    sheet_name: str | int = 0,
) -> list[Measurement]:
    if not path.exists():
        raise FileNotFoundError(f"유속 파일을 찾을 수 없습니다: {path}")

    frame = pd.read_excel(path, sheet_name=sheet_name)

    number_column = find_column(
        frame,
        {"측선번호", "측선", "번호", "no", "id", "name"},
        fallback_index=0,
    )
    x_column = find_column(
        frame,
        {
            "측선거리",
            "측선거리(m)",
            "측선위치",
            "측선위치(m)",
            "x",
            "x(m)",
            "distance",
            "station",
        },
        fallback_index=1,
    )
    velocity_column = find_column(
        frame,
        {
            "유속",
            "유속(m/s)",
            "표면유속",
            "표면유속(m/s)",
            "velocity",
            "surfacevelocity",
        },
        fallback_index=2,
    )

    weight_column = None
    for column in frame.columns:
        if normalize_header(column) in {
            normalize_header("가중치"),
            normalize_header("품질가중치"),
            normalize_header("weight"),
            normalize_header("quality_weight"),
        }:
            weight_column = column
            break

    measurements: list[Measurement] = []

    for excel_row, row in frame.iterrows():
        x = pd.to_numeric(row[x_column], errors="coerce")
        velocity = pd.to_numeric(row[velocity_column], errors="coerce")

        if pd.isna(x) or pd.isna(velocity):
            continue

        if float(velocity) < 0:
            raise ValueError(
                f"Excel {excel_row + 2}행에 음의 유속이 있습니다: "
                f"{velocity}. 초기 모델은 0 이상의 유속을 사용합니다."
            )

        raw_name = row[number_column]
        name = (
            f"V{len(measurements) + 1}"
            if pd.isna(raw_name)
            else str(raw_name).strip()
        )

        if weight_column is None or pd.isna(row[weight_column]):
            weight = 1.0
        else:
            weight = float(row[weight_column])

        if weight <= 0:
            raise ValueError(
                f"Excel {excel_row + 2}행의 가중치는 0보다 커야 합니다."
            )

        measurements.append(
            Measurement(
                x=float(x),
                surface_velocity=float(velocity),
                name=name,
                quality_weight=weight,
                source_row=int(excel_row + 2),
            )
        )

    if not measurements:
        raise ValueError("유효한 유속 측선자료가 없습니다.")

    return measurements


# =============================================================================
# 3. 횡단면 및 엔트로피 계산
# =============================================================================

def phi_from_M(M: float) -> float:
    """phi(M) = exp(M)/(exp(M)-1) - 1/M."""
    if M <= 0:
        raise ValueError("M은 0보다 커야 합니다.")
    return 1.0 + 1.0 / math.expm1(M) - 1.0 / M


def solve_M_from_phi(phi: float) -> float:
    if not 0.5 < phi < 1.0:
        raise ValueError("phi는 0.5보다 크고 1.0보다 작아야 합니다.")

    lo, hi = 1.0e-8, 100.0
    for _ in range(200):
        mid = 0.5 * (lo + hi)
        if phi_from_M(mid) < phi:
            lo = mid
        else:
            hi = mid
    return 0.5 * (lo + hi)


def crossing_x(
    x1: float,
    y1: float,
    x2: float,
    y2: float,
    water_level: float,
) -> float:
    if math.isclose(y1, y2):
        return x1
    return x1 + (water_level - y1) * (x2 - x1) / (y2 - y1)


def find_wetted_segments(
    section: np.ndarray,
    water_level: float,
) -> list[tuple[float, float]]:
    """수위선 아래에 존재하는 연속 침수구간을 찾는다."""
    segments: list[tuple[float, float]] = []
    # The section may already be wet at its first point.  Without this
    # initialization, an entirely submerged section returns no segment and a
    # section that exits the water later loses its leftmost wet portion.
    start: float | None = (
        float(section[0, 0])
        if section[0, 1] <= water_level
        else None
    )

    for idx in range(len(section) - 1):
        x1, y1 = section[idx]
        x2, y2 = section[idx + 1]
        wet1 = y1 <= water_level
        wet2 = y2 <= water_level

        if not wet1 and wet2:
            start = crossing_x(x1, y1, x2, y2, water_level)

        elif wet1 and not wet2:
            end = crossing_x(x1, y1, x2, y2, water_level)
            if start is None:
                start = x1
            segments.append((float(start), float(end)))
            start = None

    if start is not None:
        segments.append((float(start), float(section[-1, 0])))

    return segments


def select_channel_segment(
    section: np.ndarray,
    wetted_segments: list[tuple[float, float]],
) -> tuple[float, float]:
    """
    현재 수위에서 횡단면 최심점이 포함된 연속 침수구간을
    주수로로 선택한다.

    측선 위치는 주수로 선택 기준으로 사용하지 않는다. 저수위에서
    통수단면 밖에 남아 있는 과거·고정 측선 때문에 주수로 선택이
    실패하는 것을 방지하기 위함이다.
    """
    if not wetted_segments:
        raise ValueError("현재 수위에서 침수된 구간이 없습니다.")

    deepest_x = float(
        section[np.argmin(section[:, 1]), 0]
    )

    for segment in wetted_segments:
        if segment[0] <= deepest_x <= segment[1]:
            return segment

    # 부동소수점 경계 등의 예외 시 가장 넓은 침수구간을 선택
    return max(
        wetted_segments,
        key=lambda item: item[1] - item[0],
    )


def locate_wetted_segment_index(
    x: float,
    wetted_segments: list[tuple[float, float]],
) -> int | None:
    """측선 x가 포함된 침수구간 번호를 반환한다."""
    for index, segment in enumerate(wetted_segments):
        if segment[0] <= x <= segment[1]:
            return index
    return None


def build_channel_profile(
    section: np.ndarray,
    water_level: float,
    left_bank_x: float,
    right_bank_x: float,
) -> tuple[np.ndarray, np.ndarray]:
    """
    선택 수로의 하상선을 만든다.

    같은 x의 서로 다른 표고는 수직벽 또는 구조물을 나타낼 수 있으므로
    단면적 계산에서는 해당 x의 최저 표고를 사용한다.
    """
    inside = (
        (section[:, 0] >= left_bank_x)
        & (section[:, 0] <= right_bank_x)
    )
    points = section[inside]
    points = np.vstack(
        [
            [left_bank_x, water_level],
            points,
            [right_bank_x, water_level],
        ]
    )

    profile = (
        pd.DataFrame(points, columns=["x", "bed_elevation"])
        .groupby("x", as_index=False)["bed_elevation"]
        .min()
        .sort_values("x")
    )

    return (
        profile["x"].to_numpy(),
        profile["bed_elevation"].to_numpy(),
    )


def area_of_segment(
    section: np.ndarray,
    water_level: float,
    segment: tuple[float, float],
) -> tuple[float, float, float]:
    """침수구간 하나의 단면적, 수면폭, 최대수심을 계산한다."""
    left_x, right_x = segment
    profile_x, profile_bed = build_channel_profile(
        section,
        water_level,
        left_x,
        right_x,
    )

    depths = np.maximum(water_level - profile_bed, 0.0)
    area = float(np.trapezoid(depths, profile_x))
    width = float(right_x - left_x)
    max_depth = float(np.max(depths)) if len(depths) else 0.0

    return area, width, max_depth


def select_segment_by_reference_x(
    segments: list[tuple[float, float]],
    reference_x: float,
) -> tuple[float, float] | None:
    for segment in segments:
        if segment[0] <= reference_x <= segment[1]:
            return segment
    return None


def build_water_level_area_table(
    section: np.ndarray,
    config: ModelConfig,
    reference_x: float,
) -> pd.DataFrame:
    interval = config.area_table_interval_m
    if interval <= 0:
        raise ValueError("수위-단면적표 간격은 0보다 커야 합니다.")

    minimum_bed = float(np.min(section[:, 1]))

    minimum_level = (
        math.floor(minimum_bed / interval) * interval
        if config.area_table_min_level is None
        else config.area_table_min_level
    )
    maximum_level = (
        config.water_level
        if config.area_table_max_level is None
        else config.area_table_max_level
    )

    if maximum_level < minimum_level:
        raise ValueError(
            "수위-단면적표의 최대수위가 최소수위보다 작습니다."
        )

    levels = list(
        np.arange(
            minimum_level,
            maximum_level + interval * 0.5,
            interval,
        )
    )

    # 현재 수위가 간격눈금에 정확히 포함되지 않더라도 반드시 추가한다.
    if not any(
        math.isclose(level, config.water_level, abs_tol=1.0e-9)
        for level in levels
    ):
        levels.append(config.water_level)

    levels = sorted({round(float(level), 10) for level in levels})

    rows: list[dict[str, float | int]] = []

    for level in levels:
        segments = find_wetted_segments(section, level)

        total_area = 0.0
        for segment in segments:
            segment_area, _, _ = area_of_segment(
                section,
                level,
                segment,
            )
            total_area += segment_area

        main_segment = select_segment_by_reference_x(
            segments,
            reference_x,
        )

        if main_segment is None:
            main_area = 0.0
            main_width = 0.0
            main_depth = 0.0
        else:
            main_area, main_width, main_depth = area_of_segment(
                section,
                level,
                main_segment,
            )

        rows.append(
            {
                "수위(EL.m)": level,
                "주수로 단면적(m²)": main_area,
                "전체 침수단면적(m²)": total_area,
                "주수로 수면폭(m)": main_width,
                "주수로 최대수심(m)": main_depth,
                "침수구간 수": len(segments),
            }
        )

    return pd.DataFrame(rows)


def uplus_from_xi(
    xi: np.ndarray | float,
    M: float,
) -> np.ndarray | float:
    xi_array = np.asarray(xi, dtype=float)
    xi_array = np.clip(xi_array, 0.0, 1.0)
    result = np.log1p(np.expm1(M) * xi_array) / M

    if np.ndim(xi) == 0:
        return float(result)
    return result


def dimensionless_surface_velocity(
    measurement_x: np.ndarray,
    measurement_depths: np.ndarray,
    flow_center_x: float,
    left_bank_x: float,
    right_bank_x: float,
    beta_left: float,
    beta_right: float,
    M: float,
    config: ModelConfig,
) -> np.ndarray:
    if measurement_x.shape != measurement_depths.shape:
        raise ValueError(
            "measurement_x와 measurement_depths의 크기가 서로 다릅니다."
        )

    values = np.empty_like(measurement_x, dtype=float)

    for idx, (x, depth) in enumerate(
        zip(measurement_x, measurement_depths)
    ):
        if depth <= 1.0e-9:
            values[idx] = np.nan
            continue

        left_side = x <= flow_center_x

        B = (
            flow_center_x - left_bank_x + config.delta_z_left
            if left_side
            else right_bank_x - flow_center_x + config.delta_z_right
        )
        beta = beta_left if left_side else beta_right

        if B <= 0:
            raise ValueError(
                "흐름 중심과 수로 경계로 계산한 B가 0 이하입니다."
            )

        Z = abs(x - flow_center_x) / B

        if not 0.0 <= Z <= 1.0 + 1.0e-9:
            values[idx] = np.nan
            continue

        # 양의 ratio는 수면 아래 방향을 의미한다.
        # h = -ratio × 국부수심
        local_h = (
            -config.max_velocity_depth_ratio
            * depth
        )
        denominator_y = (
            depth
            + config.delta_y
            + local_h
        )

        if denominator_y <= 1.0e-9:
            values[idx] = np.nan
            continue

        # 수표면의 무차원 연직좌표
        # ratio=0.0이면 Y_surface=1.0이 되어 수면이 최대유속 위치가 된다.
        Y_surface = (
            depth + config.delta_y
        ) / denominator_y

        base = max(1.0 - Z, 0.0)

        xi = (
            Y_surface
            * (base**beta)
            * math.exp(beta * Z - Y_surface + 1.0)
        )
        values[idx] = uplus_from_xi(xi, M)

    return values


def estimate_umax(
    uplus_measured: np.ndarray,
    measured_velocity: np.ndarray,
    weights: np.ndarray,
    method: str,
) -> float:
    valid = (
        np.isfinite(uplus_measured)
        & (uplus_measured > 1.0e-8)
        & np.isfinite(measured_velocity)
    )

    if not np.any(valid):
        raise ValueError(
            "최대유속을 추정할 수 있는 유효 측선이 없습니다."
        )

    up = uplus_measured[valid]
    velocity = measured_velocity[valid]
    weight = weights[valid]
    individual = velocity / up

    if method == "least_squares":
        denominator = np.sum(weight * up * up)
        if denominator <= 0:
            raise ValueError(
                "최소제곱 최대유속 계산의 분모가 0입니다."
            )
        return float(
            np.sum(weight * up * velocity) / denominator
        )

    if method == "median":
        return float(np.median(individual))

    if method == "mean":
        return float(np.average(individual, weights=weight))

    raise ValueError(
        f"지원하지 않는 max_velocity_method: {method}"
    )


def apply_umax_constraint(
    fitted_velocity_scale: float,
    measured_velocity: np.ndarray,
    config: ModelConfig,
) -> tuple[float, float, bool]:
    """
    최소제곱 등으로 구한 속도 배율과 관측 최대표면유속을 분리한다.

    반환값
    - observed_max_surface: 관측된 최대표면유속
    - model_umax: 유속장 복원에 실제 사용할 최대유속
    - constraint_applied: 하한 제약 적용 여부
    """
    valid = np.isfinite(measured_velocity)
    if not np.any(valid):
        raise ValueError("관측 최대표면유속을 계산할 자료가 없습니다.")

    observed_max_surface = float(np.max(measured_velocity[valid]))

    if config.enforce_umax_ge_observed_surface:
        model_umax = max(
            float(fitted_velocity_scale),
            observed_max_surface,
        )
    else:
        model_umax = float(fitted_velocity_scale)

    constraint_applied = not math.isclose(
        model_umax,
        float(fitted_velocity_scale),
        rel_tol=1.0e-12,
        abs_tol=1.0e-12,
    )

    return observed_max_surface, model_umax, constraint_applied


def auto_fit_beta(
    measurement_x: np.ndarray,
    measurement_depths: np.ndarray,
    measured_velocity: np.ndarray,
    weights: np.ndarray,
    flow_center_x: float,
    left_bank_x: float,
    right_bank_x: float,
    M: float,
    config: ModelConfig,
) -> tuple[float, float, float, float, bool]:
    """
    공통 beta를 탐색한다.

    각 beta 후보에서 먼저 비제약 최소제곱 속도배율을 구한 뒤,
    설정에 따라 Umax >= 관측 최대표면유속 제약을 적용하고
    제약 적용 후 RMSE가 가장 작은 beta를 선택한다.
    """
    candidates = np.geomspace(
        config.beta_search_min,
        config.beta_search_max,
        config.beta_search_steps,
    )

    best_beta = config.beta_left
    best_fitted_scale = math.nan
    best_model_umax = math.nan
    best_rmse = math.inf
    best_constraint_applied = False

    for beta in candidates:
        uplus = dimensionless_surface_velocity(
            measurement_x,
            measurement_depths,
            flow_center_x,
            left_bank_x,
            right_bank_x,
            float(beta),
            float(beta),
            M,
            config,
        )

        # Some beta candidates can make every observation invalid.  Such a
        # candidate should not abort the entire search.
        try:
            fitted_scale = estimate_umax(
                uplus,
                measured_velocity,
                weights,
                "least_squares",
            )
        except ValueError:
            continue

        _, model_umax, constraint_applied = apply_umax_constraint(
            fitted_scale,
            measured_velocity,
            config,
        )

        predicted = model_umax * uplus
        valid = np.isfinite(predicted)
        if not np.any(valid):
            continue

        rmse = math.sqrt(
            float(
                np.average(
                    (
                        predicted[valid]
                        - measured_velocity[valid]
                    )
                    ** 2,
                    weights=weights[valid],
                )
            )
        )

        if rmse < best_rmse:
            best_beta = float(beta)
            best_fitted_scale = float(fitted_scale)
            best_model_umax = float(model_umax)
            best_rmse = float(rmse)
            best_constraint_applied = bool(constraint_applied)

    if not math.isfinite(best_rmse):
        raise ValueError(
            "No valid beta candidate produced a usable prediction."
        )

    return (
        best_beta,
        best_fitted_scale,
        best_model_umax,
        best_rmse,
        best_constraint_applied,
    )


def auto_fit_flow_center_asymmetric_beta(
    measurement_x: np.ndarray,
    measurement_depths: np.ndarray,
    measured_velocity: np.ndarray,
    weights: np.ndarray,
    left_bank_x: float,
    right_bank_x: float,
    M: float,
    config: ModelConfig,
) -> tuple[float, float, float, float, float, float, bool]:
    """
    흐름 중심 x와 좌·우 beta를 측정 표면유속에 공동 적합한다.

    각 후보 조합에서 속도배율은 가중 최소제곱으로 계산하고,
    필요하면 Umax >= 관측 최대표면유속 제약을 적용한 뒤
    제약 적용 후 가중 RMSE가 최소인 조합을 선택한다.
    """
    if config.joint_fit_center_search_steps < 3:
        raise ValueError(
            "joint_fit_center_search_steps는 3 이상이어야 합니다."
        )
    if config.joint_fit_beta_search_steps < 3:
        raise ValueError(
            "joint_fit_beta_search_steps는 3 이상이어야 합니다."
        )
    if config.joint_fit_refinement_rounds < 0:
        raise ValueError(
            "joint_fit_refinement_rounds는 0 이상이어야 합니다."
        )
    if not 0.0 < config.beta_search_min < config.beta_search_max:
        raise ValueError(
            "beta_search_min과 beta_search_max 범위를 확인해 주세요."
        )

    finite = (
        np.isfinite(measurement_x)
        & np.isfinite(measurement_depths)
        & (measurement_depths > 1.0e-9)
        & np.isfinite(measured_velocity)
        & np.isfinite(weights)
        & (weights > 0.0)
    )
    if np.count_nonzero(finite) < 4:
        raise ValueError(
            "흐름 중심과 좌·우 beta 공동 적합에는 유효 측선이 "
            "최소 4개 필요합니다."
        )

    x = measurement_x[finite]
    depths = measurement_depths[finite]
    velocity = measured_velocity[finite]
    fit_weights = weights[finite]

    if len(np.unique(x)) < 3:
        raise ValueError(
            "흐름 중심 자동 적합에는 서로 다른 측선 위치가 "
            "최소 3개 필요합니다."
        )

    channel_width = right_bank_x - left_bank_x
    margin = max(channel_width * 1.0e-6, 1.0e-6)

    # 관측 최대표면유속 측선이 실제 최대유속축에 가장 가까운 측선이라는
    # 가정 아래, 중심 탐색범위를 해당 측선의 좌·우 중간경계로 제한한다.
    # 이렇게 하면 최심점 고정은 피하면서도 비대칭 beta가 중심 위치를
    # 측정 최대값과 무관한 곳으로 과도하게 이동시키는 것을 방지한다.
    order = np.argsort(x)
    sorted_x = x[order]
    sorted_velocity = velocity[order]
    observed_max = float(np.max(sorted_velocity))
    peak_indices = np.flatnonzero(
        np.isclose(
            sorted_velocity,
            observed_max,
            rtol=1.0e-10,
            atol=1.0e-12,
        )
    )
    first_peak = int(peak_indices[0])
    last_peak = int(peak_indices[-1])

    if first_peak > 0:
        peak_cell_low = 0.5 * (
            sorted_x[first_peak - 1]
            + sorted_x[first_peak]
        )
    else:
        peak_cell_low = sorted_x[first_peak] + margin

    if last_peak < len(sorted_x) - 1:
        peak_cell_high = 0.5 * (
            sorted_x[last_peak]
            + sorted_x[last_peak + 1]
        )
    else:
        peak_cell_high = sorted_x[last_peak] - margin

    absolute_center_low = max(
        left_bank_x + margin,
        float(np.min(x)) + margin,
        float(peak_cell_low),
    )
    absolute_center_high = min(
        right_bank_x - margin,
        float(np.max(x)) - margin,
        float(peak_cell_high),
    )
    if absolute_center_high <= absolute_center_low:
        raise ValueError(
            "흐름 중심 자동 적합을 위한 유효한 중심 탐색범위가 없습니다."
        )

    absolute_log_beta_low = math.log(config.beta_search_min)
    absolute_log_beta_high = math.log(config.beta_search_max)

    center_low = absolute_center_low
    center_high = absolute_center_high
    left_log_beta_low = absolute_log_beta_low
    left_log_beta_high = absolute_log_beta_high
    right_log_beta_low = absolute_log_beta_low
    right_log_beta_high = absolute_log_beta_high

    best_center = math.nan
    best_beta_left = math.nan
    best_beta_right = math.nan
    best_fitted_scale = math.nan
    best_model_umax = math.nan
    best_rmse = math.inf
    best_constraint_applied = False

    center_steps = config.joint_fit_center_search_steps
    beta_steps = config.joint_fit_beta_search_steps

    for refinement_index in range(
        config.joint_fit_refinement_rounds + 1
    ):
        center_candidates = np.linspace(
            center_low,
            center_high,
            center_steps,
        )
        beta_left_candidates = np.exp(
            np.linspace(
                left_log_beta_low,
                left_log_beta_high,
                beta_steps,
            )
        )
        beta_right_candidates = np.exp(
            np.linspace(
                right_log_beta_low,
                right_log_beta_high,
                beta_steps,
            )
        )

        round_best_center = math.nan
        round_best_beta_left = math.nan
        round_best_beta_right = math.nan
        round_best_fitted_scale = math.nan
        round_best_model_umax = math.nan
        round_best_rmse = math.inf
        round_best_constraint_applied = False

        for center in center_candidates:
            # 좌측과 우측에 각각 적어도 한 개의 측선이 있어야 한다.
            if not (
                np.any(x < center - 1.0e-9)
                and np.any(x > center + 1.0e-9)
            ):
                continue

            for beta_left in beta_left_candidates:
                for beta_right in beta_right_candidates:
                    uplus = dimensionless_surface_velocity(
                        x,
                        depths,
                        float(center),
                        left_bank_x,
                        right_bank_x,
                        float(beta_left),
                        float(beta_right),
                        M,
                        config,
                    )

                    # 후보별 사용자료 수가 달라지면 RMSE 비교가 왜곡되므로
                    # 모든 적합 측선이 유효한 후보만 사용한다.
                    valid = np.isfinite(uplus) & (uplus > 1.0e-8)
                    if not np.all(valid):
                        continue

                    try:
                        fitted_scale = estimate_umax(
                            uplus,
                            velocity,
                            fit_weights,
                            "least_squares",
                        )
                    except ValueError:
                        continue

                    (
                        _,
                        model_umax,
                        constraint_applied,
                    ) = apply_umax_constraint(
                        fitted_scale,
                        velocity,
                        config,
                    )

                    predicted = model_umax * uplus
                    rmse = math.sqrt(
                        float(
                            np.average(
                                (predicted - velocity) ** 2,
                                weights=fit_weights,
                            )
                        )
                    )

                    if rmse < round_best_rmse:
                        round_best_center = float(center)
                        round_best_beta_left = float(beta_left)
                        round_best_beta_right = float(beta_right)
                        round_best_fitted_scale = float(fitted_scale)
                        round_best_model_umax = float(model_umax)
                        round_best_rmse = float(rmse)
                        round_best_constraint_applied = bool(
                            constraint_applied
                        )

        if not math.isfinite(round_best_rmse):
            raise ValueError(
                "흐름 중심과 좌·우 beta 공동 적합에서 "
                "유효한 후보를 찾지 못했습니다."
            )

        best_center = round_best_center
        best_beta_left = round_best_beta_left
        best_beta_right = round_best_beta_right
        best_fitted_scale = round_best_fitted_scale
        best_model_umax = round_best_model_umax
        best_rmse = round_best_rmse
        best_constraint_applied = round_best_constraint_applied

        if refinement_index == config.joint_fit_refinement_rounds:
            break

        # 현재 최적점 주변으로 탐색구간을 축소하여 재탐색한다.
        center_spacing = (
            (center_high - center_low)
            / max(center_steps - 1, 1)
        )
        left_log_spacing = (
            (left_log_beta_high - left_log_beta_low)
            / max(beta_steps - 1, 1)
        )
        right_log_spacing = (
            (right_log_beta_high - right_log_beta_low)
            / max(beta_steps - 1, 1)
        )

        center_half_span = max(
            center_spacing * 2.5,
            channel_width * 1.0e-5,
        )
        left_log_half_span = max(left_log_spacing * 2.5, 1.0e-5)
        right_log_half_span = max(right_log_spacing * 2.5, 1.0e-5)

        center_low = max(
            absolute_center_low,
            best_center - center_half_span,
        )
        center_high = min(
            absolute_center_high,
            best_center + center_half_span,
        )
        left_log_beta_low = max(
            absolute_log_beta_low,
            math.log(best_beta_left) - left_log_half_span,
        )
        left_log_beta_high = min(
            absolute_log_beta_high,
            math.log(best_beta_left) + left_log_half_span,
        )
        right_log_beta_low = max(
            absolute_log_beta_low,
            math.log(best_beta_right) - right_log_half_span,
        )
        right_log_beta_high = min(
            absolute_log_beta_high,
            math.log(best_beta_right) + right_log_half_span,
        )

    return (
        best_center,
        best_beta_left,
        best_beta_right,
        best_fitted_scale,
        best_model_umax,
        best_rmse,
        best_constraint_applied,
    )


def calculate_velocity_field(
    x_centers: np.ndarray,
    elevation_centers: np.ndarray,
    bed_at_x: np.ndarray,
    flow_center_x: float,
    left_bank_x: float,
    right_bank_x: float,
    beta_left: float,
    beta_right: float,
    umax: float,
    M: float,
    config: ModelConfig,
) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    X, EL = np.meshgrid(x_centers, elevation_centers)
    BED = np.broadcast_to(bed_at_x, X.shape)
    DEPTH = config.water_level - BED
    Y_LOCAL = EL - BED

    wet = (
        (DEPTH > 0.0)
        & (Y_LOCAL >= 0.0)
        & (EL <= config.water_level)
    )

    uplus = np.full(X.shape, np.nan, dtype=float)

    left_side = X <= flow_center_x

    B = np.where(
        left_side,
        flow_center_x - left_bank_x + config.delta_z_left,
        right_bank_x - flow_center_x + config.delta_z_right,
    )
    beta = np.where(
        left_side,
        beta_left,
        beta_right,
    )

    # 각 횡방향 위치의 국부수심에 비례하여 최대유속 발생 깊이를 설정한다.
    # 양의 ratio는 수면 아래 방향이므로 h는 음수이다.
    local_h = (
        -config.max_velocity_depth_ratio
        * DEPTH
    )
    denominator_y = (
        DEPTH
        + config.delta_y
        + local_h
    )

    valid = (
        wet
        & (denominator_y > 1.0e-9)
        & (B > 1.0e-9)
    )

    Y = np.zeros_like(X, dtype=float)
    Z = np.zeros_like(X, dtype=float)

    Y[valid] = (
        Y_LOCAL[valid] + config.delta_y
    ) / denominator_y[valid]

    Z[valid] = (
        np.abs(X[valid] - flow_center_x)
        / B[valid]
    )

    valid &= (Z >= 0.0) & (Z <= 1.0 + 1.0e-9)

    base = np.clip(1.0 - Z[valid], 0.0, 1.0)

    xi = (
        Y[valid]
        * np.power(base, beta[valid])
        * np.exp(
            beta[valid] * Z[valid]
            - Y[valid]
            + 1.0
        )
    )

    uplus[valid] = uplus_from_xi(xi, M)
    velocity = umax * uplus

    return velocity, uplus, wet


def make_edges(
    start: float,
    stop: float,
    maximum_spacing: float,
) -> np.ndarray:
    if maximum_spacing <= 0:
        raise ValueError("격자간격은 0보다 커야 합니다.")

    # Matplotlib contour/contourf requires at least a 2 x 2 grid.  Requiring
    # two cells in each direction also keeps very small sections plottable.
    count = max(
        2,
        math.ceil((stop - start) / maximum_spacing),
    )

    return np.linspace(start, stop, count + 1)


# =============================================================================
# 4. Excel 결과 저장
# =============================================================================

def write_area_table_excel(
    area_table: pd.DataFrame,
    output_path: Path,
    *,
    config: ModelConfig,
    reference_x: float,
    current_main_area: float,
    current_total_area: float,
) -> None:
    settings = pd.DataFrame(
        [
            ["횡단면 입력파일", str(config.cross_section_file)],
            ["유속 입력파일", str(config.velocity_file)],
            ["현재 수위(EL.m)", config.water_level],
            [
                "최대유속 수심비",
                config.max_velocity_depth_ratio,
            ],
            ["수위 간격(m)", config.area_table_interval_m],
            ["주수로 기준 위치 x(m)", reference_x],
            ["현재수위 주수로 단면적(m²)", current_main_area],
            ["현재수위 전체 침수단면적(m²)", current_total_area],
            [
                "비고",
                (
                    "주수로 단면적은 기준 위치 x가 포함된 연속 "
                    "침수구간을 대상으로 계산"
                ),
            ],
        ],
        columns=["항목", "값"],
    )

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:
        area_table.to_excel(
            writer,
            sheet_name="수위-단면적",
            index=False,
        )
        settings.to_excel(
            writer,
            sheet_name="설정",
            index=False,
        )

    workbook = load_workbook(output_path)
    table_sheet = workbook["수위-단면적"]
    setting_sheet = workbook["설정"]

    header_fill = PatternFill(
        "solid",
        fgColor="0F6B78",
    )
    header_font = Font(
        bold=True,
        color="FFFFFF",
    )
    sub_fill = PatternFill(
        "solid",
        fgColor="DCEFF2",
    )
    current_fill = PatternFill(
        "solid",
        fgColor="FFF2CC",
    )
    thin_gray = Side(
        style="thin",
        color="D9E1F2",
    )

    for sheet in (table_sheet, setting_sheet):
        sheet.freeze_panes = "A2"

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(
                    left=thin_gray,
                    right=thin_gray,
                    top=thin_gray,
                    bottom=thin_gray,
                )

    # 현재 수위 행 강조
    for row in range(2, table_sheet.max_row + 1):
        value = table_sheet.cell(row=row, column=1).value
        if (
            isinstance(value, (int, float))
            and math.isclose(
                float(value),
                config.water_level,
                abs_tol=1.0e-9,
            )
        ):
            for column in range(1, table_sheet.max_column + 1):
                table_sheet.cell(row=row, column=column).fill = current_fill
            break

    for row in range(2, table_sheet.max_row + 1):
        table_sheet.cell(row=row, column=1).number_format = "0.00"
        for column in range(2, 6):
            table_sheet.cell(
                row=row,
                column=column,
            ).number_format = "0.000"
        table_sheet.cell(
            row=row,
            column=6,
        ).number_format = "0"

    # 설정 시트의 항목열 강조
    for row in range(2, setting_sheet.max_row + 1):
        setting_sheet.cell(row=row, column=1).fill = sub_fill
        setting_sheet.cell(row=row, column=1).font = Font(bold=True)

    setting_sheet.column_dimensions["A"].width = 30
    setting_sheet.column_dimensions["B"].width = 78
    setting_sheet["B10"].alignment = Alignment(
        wrap_text=True,
        vertical="top",
    )

    widths = [14, 21, 24, 20, 22, 13]
    for index, width in enumerate(widths, start=1):
        table_sheet.column_dimensions[
            get_column_letter(index)
        ].width = width

    table_sheet.auto_filter.ref = table_sheet.dimensions

    chart = ScatterChart()
    chart.title = "수위-주수로 단면적 관계"
    chart.style = 13
    chart.y_axis.title = "수위 (EL.m)"
    chart.x_axis.title = "단면적 (m²)"
    chart.height = 10
    chart.width = 18
    chart.scatterStyle = "line"

    x_values = Reference(
        table_sheet,
        min_col=1,
        min_row=2,
        max_row=table_sheet.max_row,
    )
    y_values = Reference(
        table_sheet,
        min_col=2,
        min_row=2,
        max_row=table_sheet.max_row,
    )
    series = Series(
        x_values,
        y_values,
        title="주수로 단면적(m²)",
    )
    series.graphicalProperties.line.width = 22000
    series.marker.symbol = "none"
    chart.series.append(series)
    table_sheet.add_chart(chart, "H2")

    workbook.save(output_path)


# =============================================================================
# 5. 전체 계산
# =============================================================================

def run_model(
    section: np.ndarray,
    measurements: list[Measurement],
    config: ModelConfig,
) -> dict:
    if not 0.0 <= config.max_velocity_depth_ratio < 1.0:
        raise ValueError(
            "max_velocity_depth_ratio는 0.0 이상 1.0 미만이어야 합니다."
        )

    if config.outside_wetted_measurement_policy not in {"exclude", "error"}:
        raise ValueError(
            "outside_wetted_measurement_policy는 "
            "'exclude' 또는 'error'여야 합니다."
        )

    if config.minimum_positive_measurements < 1:
        raise ValueError(
            "minimum_positive_measurements는 1 이상이어야 합니다."
        )

    if config.contour_level_count < 2:
        raise ValueError("contour_level_count must be at least 2.")

    if any(m.quality_weight <= 0 for m in measurements):
        raise ValueError(
            "quality_weight는 모두 0보다 커야 합니다."
        )

    wetted_segments = find_wetted_segments(
        section,
        config.water_level,
    )

    if not wetted_segments:
        raise ValueError(
            "현재 수위에서 침수된 횡단면 구간이 없습니다."
        )

    left_bank_x, right_bank_x = select_channel_segment(
        section,
        wetted_segments,
    )

    profile_x, profile_bed = build_channel_profile(
        section,
        config.water_level,
        left_bank_x,
        right_bank_x,
    )

    exact_area = float(
        np.trapezoid(
            config.water_level - profile_bed,
            profile_x,
        )
    )

    flow_center_x_was_user_specified = config.flow_center_x is not None
    if config.flow_center_x is None:
        flow_center_x = float(
            profile_x[np.argmin(profile_bed)]
        )
        flow_center_source = "최심점 초기값"
    else:
        flow_center_x = float(config.flow_center_x)
        flow_center_source = "사용자 지정"

    if not left_bank_x < flow_center_x < right_bank_x:
        raise ValueError(
            "flow_center_x는 선택 수로 내부에 있어야 합니다."
        )

    M = (
        solve_M_from_phi(config.phi)
        if config.entropy_M is None
        else float(config.entropy_M)
    )
    phi_used = phi_from_M(M)

    # 현재 수위에서 각 측선이 어느 침수구간에 있는지 판정한다.
    measurement_segment_indices = [
        locate_wetted_segment_index(
            measurement.x,
            wetted_segments,
        )
        for measurement in measurements
    ]

    inside_channel = np.asarray(
        [
            left_bank_x <= measurement.x <= right_bank_x
            for measurement in measurements
        ],
        dtype=bool,
    )

    excluded_outside_measurements = [
        measurement
        for measurement, inside in zip(
            measurements,
            inside_channel,
        )
        if not inside
    ]
    excluded_outside_positive = [
        measurement
        for measurement in excluded_outside_measurements
        if measurement.surface_velocity > 0.0
    ]

    if (
        excluded_outside_positive
        and config.outside_wetted_measurement_policy == "error"
    ):
        outside_text = ", ".join(
            f"{measurement.name}(x={measurement.x:.2f} m, "
            f"V={measurement.surface_velocity:.3f} m/s)"
            for measurement in excluded_outside_positive
        )
        raise ValueError(
            "현재 수위의 선택 주수로 밖에 양의 유속 측선이 있습니다: "
            f"{outside_text}"
        )

    if excluded_outside_measurements:
        print(
            "[경고] 현재 수위에서 선택된 주수로 밖의 측선 "
            f"{len(excluded_outside_measurements)}개를 계산에서 제외합니다."
        )
        print(
            "       선택 주수로: "
            f"x={left_bank_x:.3f} ~ {right_bank_x:.3f} m"
        )
        for measurement, segment_index in zip(
            measurements,
            measurement_segment_indices,
        ):
            if left_bank_x <= measurement.x <= right_bank_x:
                continue

            location_reason = (
                "현재 수위에서 비침수 위치"
                if segment_index is None
                else "선택 주수로와 분리된 다른 침수구간"
            )
            print(
                "       - "
                f"{measurement.name}: x={measurement.x:.3f} m, "
                f"V={measurement.surface_velocity:.3f} m/s "
                f"({location_reason})"
            )

    nonnegative_condition = np.asarray(
        [
            (
                measurement.surface_velocity >= 0.0
                if config.include_zero_velocity_in_fit
                else measurement.surface_velocity > 0.0
            )
            for measurement in measurements
        ],
        dtype=bool,
    )

    # 선택 주수로 안의 자료만 모형 적합에 사용한다.
    used_for_fit = inside_channel & nonnegative_condition

    fit_measurements = [
        measurement
        for measurement, use in zip(
            measurements,
            used_for_fit,
        )
        if use
    ]

    positive_fit_measurements = [
        measurement
        for measurement in fit_measurements
        if measurement.surface_velocity > 0.0
    ]

    if (
        len(positive_fit_measurements)
        < config.minimum_positive_measurements
    ):
        raise ValueError(
            "현재 수위의 선택 주수로 안에 양의 유속 측선이 부족합니다. "
            f"필요={config.minimum_positive_measurements}개, "
            f"현재={len(positive_fit_measurements)}개, "
            f"주수로=x {left_bank_x:.3f}~{right_bank_x:.3f} m"
        )


    measurement_x = np.asarray(
        [m.x for m in fit_measurements],
        dtype=float,
    )
    measurement_bed_elevation = np.interp(
        measurement_x,
        profile_x,
        profile_bed,
    )
    measurement_depths = (
        config.water_level
        - measurement_bed_elevation
    )

    measured_velocity = np.asarray(
        [m.surface_velocity for m in fit_measurements],
        dtype=float,
    )
    weights = np.asarray(
        [m.quality_weight for m in fit_measurements],
        dtype=float,
    )

    # Positive observations are not necessarily usable observations: zero
    # depth and invalid normalized coordinates produce NaN uplus values.
    prefit_uplus = dimensionless_surface_velocity(
        measurement_x,
        measurement_depths,
        flow_center_x,
        left_bank_x,
        right_bank_x,
        config.beta_left,
        config.beta_right,
        M,
        config,
    )
    valid_positive_measurements = (
        np.isfinite(prefit_uplus)
        & (prefit_uplus > 1.0e-8)
        & (measured_velocity > 0.0)
    )
    valid_positive_count = int(np.count_nonzero(valid_positive_measurements))
    if valid_positive_count < config.minimum_positive_measurements:
        raise ValueError(
            "Too few positive measurements remain after depth and "
            "normalized-coordinate validation: "
            f"required={config.minimum_positive_measurements}, "
            f"valid={valid_positive_count}."
        )

    observed_max_surface = float(np.max(measured_velocity))

    joint_spatial_fit_applied = False
    common_beta_fit_applied = False
    joint_fit_fallback_reason = ""

    if (
        config.auto_fit_flow_center_asymmetric_beta
        and not flow_center_x_was_user_specified
        and len(fit_measurements) >= 4
        and len(np.unique(measurement_x)) >= 3
    ):
        (
            flow_center_x,
            beta_left,
            beta_right,
            fitted_velocity_scale,
            model_umax,
            beta_fit_rmse,
            umax_constraint_applied,
        ) = auto_fit_flow_center_asymmetric_beta(
            measurement_x,
            measurement_depths,
            measured_velocity,
            weights,
            left_bank_x,
            right_bank_x,
            M,
            config,
        )
        joint_spatial_fit_applied = True
        flow_center_source = "표면유속 공동 적합"

    elif (
        config.auto_fit_common_beta
        and len(fit_measurements) >= 2
    ):
        if (
            config.auto_fit_flow_center_asymmetric_beta
            and not flow_center_x_was_user_specified
        ):
            joint_fit_fallback_reason = (
                "공동 적합에 필요한 유효 측선 수 또는 서로 다른 "
                "측선 위치가 부족하여 기존 공통 beta 적합을 사용했습니다."
            )
            print(f"[경고] {joint_fit_fallback_reason}")

        (
            common_beta,
            fitted_velocity_scale,
            model_umax,
            beta_fit_rmse,
            umax_constraint_applied,
        ) = auto_fit_beta(
            measurement_x,
            measurement_depths,
            measured_velocity,
            weights,
            flow_center_x,
            left_bank_x,
            right_bank_x,
            M,
            config,
        )
        beta_left = common_beta
        beta_right = common_beta
        common_beta_fit_applied = True

    else:
        beta_left = config.beta_left
        beta_right = config.beta_right
        beta_fit_rmse = math.nan

    uplus_measured = dimensionless_surface_velocity(
        measurement_x,
        measurement_depths,
        flow_center_x,
        left_bank_x,
        right_bank_x,
        beta_left,
        beta_right,
        M,
        config,
    )

    if not joint_spatial_fit_applied and not common_beta_fit_applied:
        fitted_velocity_scale = estimate_umax(
            uplus_measured,
            measured_velocity,
            weights,
            config.max_velocity_method,
        )
        (
            observed_max_surface,
            model_umax,
            umax_constraint_applied,
        ) = apply_umax_constraint(
            fitted_velocity_scale,
            measured_velocity,
            config,
        )

    predicted_surface_unconstrained = (
        fitted_velocity_scale * uplus_measured
    )
    residual_unconstrained = (
        measured_velocity - predicted_surface_unconstrained
    )

    predicted_surface = model_umax * uplus_measured
    residual = measured_velocity - predicted_surface

    individual_umax = np.divide(
        measured_velocity,
        uplus_measured,
        out=np.full_like(
            measured_velocity,
            np.nan,
            dtype=float,
        ),
        where=np.isfinite(uplus_measured)
        & (uplus_measured > 1.0e-8),
    )

    valid_prediction = np.isfinite(predicted_surface)
    surface_rmse = math.sqrt(
        float(
            np.average(
                residual[valid_prediction] ** 2,
                weights=weights[valid_prediction],
            )
        )
    )

    valid_unconstrained = np.isfinite(
        predicted_surface_unconstrained
    )
    surface_rmse_unconstrained = math.sqrt(
        float(
            np.average(
                residual_unconstrained[valid_unconstrained] ** 2,
                weights=weights[valid_unconstrained],
            )
        )
    )

    x_edges = make_edges(
        left_bank_x,
        right_bank_x,
        config.horizontal_grid_m,
    )
    x_centers = 0.5 * (
        x_edges[:-1] + x_edges[1:]
    )
    dx = np.diff(x_edges)

    minimum_bed = float(np.min(profile_bed))

    elevation_edges = make_edges(
        minimum_bed,
        config.water_level,
        config.vertical_grid_m,
    )
    elevation_centers = 0.5 * (
        elevation_edges[:-1] + elevation_edges[1:]
    )
    dy = np.diff(elevation_edges)

    bed_at_x = np.interp(
        x_centers,
        profile_x,
        profile_bed,
    )

    velocity, uplus_grid, _ = calculate_velocity_field(
        x_centers,
        elevation_centers,
        bed_at_x,
        flow_center_x,
        left_bank_x,
        right_bank_x,
        beta_left,
        beta_right,
        model_umax,
        M,
        config,
    )

    cell_area = dy[:, None] * dx[None, :]
    valid_velocity = np.isfinite(velocity)

    if not np.any(valid_velocity):
        raise ValueError(
            "The generated grid contains no valid wet velocity cells. "
            "Reduce the grid spacing or check the section and water level."
        )

    grid_area = float(
        np.sum(cell_area[valid_velocity])
    )
    discharge_grid = float(
        np.nansum(velocity * cell_area)
    )
    mean_velocity_grid = (
        discharge_grid / grid_area
        if grid_area > 0
        else math.nan
    )
    grid_area_relative_error = (
        abs(grid_area - exact_area) / exact_area
        if exact_area > 0.0
        else math.nan
    )

    if grid_area_relative_error > 0.05:
        print(
            "[WARNING] Grid area differs from the exact cross-section area "
            f"by {grid_area_relative_error:.1%}. Consider reducing the grid "
            "spacing."
        )

    discharge_simple_observed_max = (
        exact_area
        * phi_used
        * observed_max_surface
    )
    discharge_simple_model_umax = (
        exact_area
        * phi_used
        * model_umax
    )

    output_dir = (
        Path(__file__).resolve().parent
        / config.output_directory
    )
    output_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    # 전체 측선 결과표
    fit_lookup = {
        id(measurement): index
        for index, measurement in enumerate(
            fit_measurements
        )
    }

    fit_rows: list[dict[str, object]] = []

    for measurement, use, segment_index in zip(
        measurements,
        used_for_fit,
        measurement_segment_indices,
    ):
        if use:
            fit_index = fit_lookup[id(measurement)]
            fit_rows.append(
                {
                    "name": measurement.name,
                    "source_excel_row": measurement.source_row,
                    "x_m": measurement.x,
                    "local_water_depth_m":
                        measurement_depths[fit_index],
                    "maximum_velocity_depth_below_surface_m":
                        (
                            config.max_velocity_depth_ratio
                            * measurement_depths[fit_index]
                        ),
                    "measured_surface_velocity_mps":
                        measurement.surface_velocity,
                    "used_for_fit": True,
                    "exclusion_reason": "",
                    "dimensionless_surface_velocity":
                        uplus_measured[fit_index],
                    "individual_umax_estimate_mps":
                        individual_umax[fit_index],
                    "predicted_surface_unconstrained_mps":
                        predicted_surface_unconstrained[fit_index],
                    "residual_unconstrained_mps":
                        residual_unconstrained[fit_index],
                    "predicted_surface_velocity_mps":
                        predicted_surface[fit_index],
                    "residual_mps": residual[fit_index],
                    "quality_weight":
                        measurement.quality_weight,
                }
            )
        else:
            if not (
                left_bank_x
                <= measurement.x
                <= right_bank_x
            ):
                reason = (
                    "현재 수위에서 비침수 위치"
                    if segment_index is None
                    else "선택 주수로와 분리된 다른 침수구간"
                )
            else:
                reason = "0 m/s 자료 적합 제외 설정"

            fit_rows.append(
                {
                    "name": measurement.name,
                    "source_excel_row": measurement.source_row,
                    "x_m": measurement.x,
                    "local_water_depth_m": np.nan,
                    "maximum_velocity_depth_below_surface_m":
                        np.nan,
                    "measured_surface_velocity_mps":
                        measurement.surface_velocity,
                    "used_for_fit": False,
                    "exclusion_reason": reason,
                    "dimensionless_surface_velocity":
                        np.nan,
                    "individual_umax_estimate_mps":
                        np.nan,
                    "predicted_surface_unconstrained_mps":
                        np.nan,
                    "residual_unconstrained_mps": np.nan,
                    "predicted_surface_velocity_mps":
                        np.nan,
                    "residual_mps": np.nan,
                    "quality_weight":
                        measurement.quality_weight,
                }
            )

    fit_frame = pd.DataFrame(fit_rows)
    fit_csv = output_dir / "measurement_fit.csv"
    fit_frame.to_csv(
        fit_csv,
        index=False,
        encoding="utf-8-sig",
    )

    section_frame = pd.DataFrame(
        {
            "x_m": profile_x,
            "bed_elevation_m": profile_bed,
        }
    )
    section_csv = (
        output_dir
        / "selected_channel_cross_section.csv"
    )
    section_frame.to_csv(
        section_csv,
        index=False,
        encoding="utf-8-sig",
    )

    X, EL = np.meshgrid(
        x_centers,
        elevation_centers,
    )
    BED = np.broadcast_to(
        bed_at_x,
        X.shape,
    )

    grid_frame = pd.DataFrame(
        {
            "x_m": X[valid_velocity],
            "elevation_m": EL[valid_velocity],
            "bed_elevation_m": BED[valid_velocity],
            "water_depth_at_x_m":
                config.water_level
                - BED[valid_velocity],
            "u_plus": uplus_grid[valid_velocity],
            "velocity_mps": velocity[valid_velocity],
        }
    )
    grid_csv = output_dir / "velocity_grid.csv"
    grid_frame.to_csv(
        grid_csv,
        index=False,
        encoding="utf-8-sig",
    )

    # 수위-단면적표
    if config.area_reference_x is None:
        reference_x = float(
            section[np.argmin(section[:, 1]), 0]
        )
    else:
        reference_x = config.area_reference_x

    area_table = build_water_level_area_table(
        section,
        config,
        reference_x,
    )

    current_area_row = area_table.loc[
        np.isclose(
            area_table["수위(EL.m)"],
            config.water_level,
        )
    ]

    if current_area_row.empty:
        raise RuntimeError(
            "수위-단면적표에서 현재 수위를 찾지 못했습니다."
        )

    current_main_area = float(
        current_area_row.iloc[0][
            "주수로 단면적(m²)"
        ]
    )
    current_total_area = float(
        current_area_row.iloc[0][
            "전체 침수단면적(m²)"
        ]
    )

    area_excel = (
        output_dir
        / "water_level_area_table.xlsx"
    )
    write_area_table_excel(
        area_table,
        area_excel,
        config=config,
        reference_x=reference_x,
        current_main_area=current_main_area,
        current_total_area=current_total_area,
    )

    summary = {
        "model_status":
            "initial_prototype_not_field_calibrated",
        "cross_section_file":
            str(config.cross_section_file),
        "velocity_file":
            str(config.velocity_file),
        "water_level_m": config.water_level,
        "measurement_count_total":
            len(measurements),
        "measurement_count_used":
            len(fit_measurements),
        "measurement_count_positive_used":
            len(positive_fit_measurements),
        "measurement_count_excluded_outside_selected_channel":
            len(excluded_outside_measurements),
        "outside_wetted_measurement_policy":
            config.outside_wetted_measurement_policy,
        "minimum_positive_measurements":
            config.minimum_positive_measurements,
        "wetted_segments": [
            {
                "segment_number": index + 1,
                "left_x_m": segment[0],
                "right_x_m": segment[1],
                "width_m": segment[1] - segment[0],
            }
            for index, segment in enumerate(wetted_segments)
        ],
        "excluded_outside_measurements": [
            {
                "name": measurement.name,
                "x_m": measurement.x,
                "surface_velocity_mps": measurement.surface_velocity,
                "reason": (
                    "현재 수위에서 비침수 위치"
                    if segment_index is None
                    else "선택 주수로와 분리된 다른 침수구간"
                ),
            }
            for measurement, inside, segment_index in zip(
                measurements,
                inside_channel,
                measurement_segment_indices,
            )
            if not inside
        ],
        "selected_channel_left_x_m":
            left_bank_x,
        "selected_channel_right_x_m":
            right_bank_x,
        "selected_channel_width_m":
            right_bank_x - left_bank_x,
        "flow_center_x_m": flow_center_x,
        "flow_center_source": flow_center_source,
        "minimum_bed_elevation_m":
            minimum_bed,
        "maximum_depth_m":
            config.water_level - minimum_bed,
        "max_velocity_depth_ratio":
            config.max_velocity_depth_ratio,
        "max_velocity_depth_at_deepest_point_m":
            (
                config.max_velocity_depth_ratio
                * (config.water_level - minimum_bed)
            ),
        "exact_cross_section_area_m2":
            exact_area,
        "area_table_current_main_area_m2":
            current_main_area,
        "area_table_current_total_area_m2":
            current_total_area,
        "area_table_reference_x_m":
            reference_x,
        "grid_cross_section_area_m2":
            grid_area,
        "grid_area_relative_error":
            grid_area_relative_error,
        "phi": phi_used,
        "entropy_M": M,
        "beta_left": beta_left,
        "beta_right": beta_right,
        "auto_fit_common_beta":
            config.auto_fit_common_beta,
        "auto_fit_flow_center_asymmetric_beta":
            config.auto_fit_flow_center_asymmetric_beta,
        "joint_spatial_fit_applied":
            joint_spatial_fit_applied,
        "common_beta_fit_applied":
            common_beta_fit_applied,
        "joint_fit_fallback_reason":
            joint_fit_fallback_reason,
        "beta_fit_rmse_mps":
            beta_fit_rmse,
        "surface_fit_rmse_unconstrained_mps":
            surface_rmse_unconstrained,
        "surface_fit_rmse_mps":
            surface_rmse,
        "observed_max_surface_velocity_mps":
            observed_max_surface,
        "fitted_velocity_scale_mps":
            fitted_velocity_scale,
        "model_umax_used_mps":
            model_umax,
        "umax_constraint_enabled":
            config.enforce_umax_ge_observed_surface,
        "umax_constraint_applied":
            umax_constraint_applied,
        # 이전 결과파일과의 호환을 위한 별칭
        "estimated_umax_mps":
            model_umax,
        "grid_mean_velocity_mps":
            mean_velocity_grid,
        "grid_integrated_discharge_m3s":
            discharge_grid,
        "simple_Q_using_observed_max_surface_m3s":
            discharge_simple_observed_max,
        "simple_Q_using_model_umax_m3s":
            discharge_simple_model_umax,
        "horizontal_grid_m":
            float(np.max(dx)),
        "vertical_grid_m":
            float(np.max(dy)),
        "note": (
            "단면 내부 유속은 표면유속과 엔트로피 "
            "유속분포 모형으로부터 추정한 값이며, "
            "현장 기준유량을 이용한 검정 전 결과입니다."
        ),
    }

    summary_json = (
        output_dir
        / "calculation_summary.json"
    )
    summary_json.write_text(
        json.dumps(
            summary,
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    # 등유속선도
    fig, ax = plt.subplots(
        figsize=(16, 7),
    )

    masked_velocity = np.ma.masked_invalid(
        velocity
    )

    positive_max = max(
        float(np.nanmax(velocity)),
        0.01,
    )
    levels = np.linspace(
        0.0,
        positive_max,
        config.contour_level_count,
    )

    contour_filled = ax.contourf(
        X,
        EL,
        masked_velocity,
        levels=levels,
        extend="max",
    )

    contour_lines = ax.contour(
        X,
        EL,
        masked_velocity,
        levels=levels[1:],
        linewidths=0.55,
    )
    ax.clabel(
        contour_lines,
        inline=True,
        fontsize=7,
        fmt="%.2f",
    )

    plot_x = section[:, 0]
    plot_bed = section[:, 1]

    plot_bottom = float(np.min(plot_bed)) - 1.0
    plot_top = (
            max(
            float(np.max(plot_bed)),
            config.water_level,
        )
        + 0.5
    )

    ax.fill_between(
        plot_x,
        plot_bottom,
        plot_bed,
        alpha=0.35,
        label="River bed",
    )

    ax.plot(
        plot_x,
        plot_bed,
        linewidth=1.2,
        label="Bed elevation",
    )

    ax.hlines(
        config.water_level,
        left_bank_x,
        right_bank_x,
        linestyle="--",
        linewidth=1.2,
        label=(
            f"Water level EL. "
            f"{config.water_level:.2f} m"
        ),
    )
    for measurement in measurements:
        if not (
            left_bank_x
            <= measurement.x
            <= right_bank_x
        ):
            continue

        marker = (
            "o"
            if measurement.surface_velocity > 0
            else "x"
        )

        ax.scatter(
            measurement.x,
            config.water_level,
            s=52,
            marker=marker,
            edgecolors=(
                "black"
                if marker == "o"
                else None
            ),
            zorder=6,
        )

        ax.annotate(
            (
                f"{measurement.name}\n"
                f"{measurement.surface_velocity:.2f} m/s"
            ),
            xy=(
                measurement.x,
                config.water_level,
            ),
            xytext=(0, 10),
            textcoords="offset points",
            ha="center",
            va="bottom",
            fontsize=7,
        )

    ax.axvline(
        flow_center_x,
        linestyle=":",
        linewidth=1.0,
        label=(
            f"Flow center x="
            f"{flow_center_x:.2f} m"
        ),
    )

    colorbar = fig.colorbar(
        contour_filled,
        ax=ax,
        pad=0.01,
    )
    colorbar.set_label(
        "Estimated velocity (m/s)"
    )

    ax.set_title(
        "Entropy-based Estimated Isovelocity Contour\n"
        f"Q={discharge_grid:,.2f} m³/s, "
        f"Mean velocity={mean_velocity_grid:.3f} m/s, "
        #f"Observed Vmax={observed_max_surface:.3f} m/s, "
        f"Model Umax={model_umax:.3f} m/s\n"
        #f"Unconstrained scale={fitted_velocity_scale:.3f} m/s, "
        f"Depth ratio={config.max_velocity_depth_ratio:.3f}, "
        f"phi={phi_used:.3f}, M={M:.3f}, "
        #f"beta={beta_left:.3f}"
    )
    ax.set_xlabel(
        "Cross-section distance x (m)"
    )
    ax.set_ylabel(
        "Elevation EL. (m)"
    )
    ax.set_xlim(
        float(np.min(plot_x)),
        float(np.max(plot_x)),
    )
    ax.set_ylim(
        plot_bottom,
        plot_top,
    )
    ax.grid(alpha=0.15)
    ax.legend(
        loc="lower right",
        fontsize=8,
    )

    ax.text(
        0.01,
        0.02,
        (
            "Internal velocities are model estimates "
            "derived from measured surface velocities."
        ),
        transform=ax.transAxes,
        fontsize=8,
        bbox={
            "facecolor": "white",
            "alpha": 0.75,
            "edgecolor": "none",
        },
    )

    figure_path = (
        output_dir
        / "isovelocity_contour.png"
    )

    fig.tight_layout()
    fig.savefig(
        figure_path,
        dpi=180,
    )

    # 파일 저장 후 화면 표시
    if config.show_plot:
        plt.show()

    plt.close(fig)

    summary["output_files"] = {
        "isovelocity_contour":
            str(figure_path),
        "water_level_area_table":
            str(area_excel),
        "velocity_grid_csv":
            str(grid_csv),
        "measurement_fit_csv":
            str(fit_csv),
        "selected_channel_cross_section_csv":
            str(section_csv),
        "summary_json":
            str(summary_json),
    }

    # Rewrite after output_files has been added so the returned dictionary and
    # the persisted JSON contain the same information.
    summary_json.write_text(
        json.dumps(
            summary,
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    return summary


# =============================================================================
# 6. 실행
# =============================================================================

def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "외부 Excel 입력 기반 엔트로피 "
            "유속분포법 계산"
        )
    )
    parser.add_argument(
        "--xs",
        type=Path,
        default=CONFIG.cross_section_file,
        help="횡단면 Excel 경로",
    )
    parser.add_argument(
        "--vel",
        type=Path,
        default=CONFIG.velocity_file,
        help="표면유속 Excel 경로",
    )
    parser.add_argument(
        "--water-level",
        type=float,
        default=CONFIG.water_level,
        help="계산 수위(EL.m)",
    )
    parser.add_argument(
        "--no-show",
        action="store_true",
        help="등유속선도 화면 표시 생략",
    )
    return parser.parse_args()


def main() -> None:
    arguments = parse_arguments()

    config = replace(
        CONFIG,
        cross_section_file=arguments.xs.resolve(),
        velocity_file=arguments.vel.resolve(),
        water_level=arguments.water_level,
        show_plot=not arguments.no_show,
    )

    section = load_cross_section_excel(
        config.cross_section_file,
        config.cross_section_sheet,
    )
    measurements = load_velocity_excel(
        config.velocity_file,
        config.velocity_sheet,
    )

    summary = run_model(
        section,
        measurements,
        config,
    )

    print("=" * 76)
    print("엔트로피 기반 유속분포법 계산 완료")
    print("=" * 76)
    print(
        f"횡단면 입력             : "
        f"{summary['cross_section_file']}"
    )
    print(
        f"유속 입력               : "
        f"{summary['velocity_file']}"
    )
    print(
        f"전체/사용 측선 수       : "
        f"{summary['measurement_count_total']} / "
        f"{summary['measurement_count_used']}"
    )
    print(
        f"사용 양의 유속 측선 수  : "
        f"{summary['measurement_count_positive_used']}"
    )
    print(
        f"주수로 밖 제외 측선 수  : "
        f"{summary['measurement_count_excluded_outside_selected_channel']}"
    )
    print(
        "선택 수로 범위          : "
        f"{summary['selected_channel_left_x_m']:.3f} ~ "
        f"{summary['selected_channel_right_x_m']:.3f} m"
    )
    print(
        f"통수단면적              : "
        f"{summary['exact_cross_section_area_m2']:.3f} m²"
    )
    print(
        f"흐름 중심 x             : "
        f"{summary['flow_center_x_m']:.3f} m"
    )
    print(
        f"phi / M                 : "
        f"{summary['phi']:.6f} / "
        f"{summary['entropy_M']:.6f}"
    )
    print(
        f"beta(L/R)               : "
        f"{summary['beta_left']:.6f} / "
        f"{summary['beta_right']:.6f}"
    )
    print(
        f"최대유속 수심비         : "
        f"{summary['max_velocity_depth_ratio']:.4f} "
        f"(최심부 기준 "
        f"{summary['max_velocity_depth_at_deepest_point_m']:.3f} m)"
    )
    print(
        f"관측 최대표면유속       : "
        f"{summary['observed_max_surface_velocity_mps']:.4f} m/s"
    )
    print(
        f"비제약 적합 속도배율    : "
        f"{summary['fitted_velocity_scale_mps']:.4f} m/s"
    )
    print(
        f"유속장 적용 Model Umax  : "
        f"{summary['model_umax_used_mps']:.4f} m/s "
        f"(제약 적용={summary['umax_constraint_applied']})"
    )
    print(
        f"단면평균유속(격자적분)   : "
        f"{summary['grid_mean_velocity_mps']:.4f} m/s"
    )
    print(
        f"유량(2차원 격자적분)     : "
        f"{summary['grid_integrated_discharge_m3s']:.3f} m³/s"
    )
    print(
        f"표면유속 적합 RMSE       : "
        f"{summary['surface_fit_rmse_mps']:.4f} m/s"
    )
    print("-" * 76)

    for name, path in summary[
        "output_files"
    ].items():
        print(f"{name:36s}: {path}")

    print("=" * 76)
    print(
        "주의: 현장 기준유량으로 검정하기 전의 "
        "시험용 결과입니다."
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(
            f"\n[오류] {type(error).__name__}: {error}",
            file=sys.stderr,
        )
        raise
